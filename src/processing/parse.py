"""Parse and normalise the RBNZ Bank Financial Strength Dashboard XLSX.

Processing layer responsibility: read raw files, apply metric mappings from
``config/metrics.yaml``, and produce canonical rows in the schema:

    entity  — institution name (e.g. "ANZ")
    metric  — canonical metric name from glossary (e.g. "CET1 Ratio")
    value   — numeric value as-is from source, or None if missing
    period  — quarter string derived from quarter-end date (e.g. "2024-Q1")
    source  — always "rbnz-dashboard" for this source

See ADR-0004 for the full data contract.
"""

from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl  # type: ignore[import-untyped]

from src.logger import get_logger

logger = get_logger(__name__)

# Row indices (0-based) in the RBNZ XLSX Data sheet
_ROW_CATEGORIES = 0
_ROW_SERIES_NAMES = 1
_ROW_NOTES = 2
_ROW_UNITS = 3
_ROW_SERIES_IDS = 4
_ROW_DATA_START = 5

# Column indices (0-based)
_COL_DATE = 0
_COL_INSTITUTION = 1
_COL_DATA_START = 2

_SOURCE_ID = "rbnz-dashboard"

# Classification of RBNZ-registered entities. Group entities have a different
# regulatory reporting boundary than their standalone subsidiary.
_GROUP_ENTITIES: frozenset[str] = frozenset(
    {
        "ANZ Group",
        "BOC Group",
        "CBA Group",
        "CCB Group",
        "ICBC Group",
        "Rabo Group",
        "WBC Group",
    }
)


def parse_rbnz_xlsx(
    xlsx_path: Path,
    metrics_config: dict[str, Any],
) -> list[dict[str, Any]]:
    """Parse the RBNZ XLSX and return rows in the canonical schema.

    Parameters
    ----------
    xlsx_path:
        Path to the RBNZ Bank Financial Strength Dashboard XLSX file.
    metrics_config:
        Loaded ``config/metrics.yaml`` content.  The function reads
        ``mappings.rbnz-dashboard`` to determine which series to extract.

    Returns
    -------
    list[dict]
        Canonical rows with keys: entity, metric, value, period, source.

    Raises
    ------
    FileNotFoundError
        If ``xlsx_path`` does not exist.
    ValueError
        If the XLSX Data sheet is missing or has too few header rows.
    """
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX not found: {xlsx_path}")

    mappings: dict[str, str] = (metrics_config.get("mappings") or {}).get(_SOURCE_ID) or {}
    if not mappings:
        logger.warning("No %s mappings found in metrics config; output will be empty", _SOURCE_ID)
        return []

    logger.info("Opening XLSX: %s", xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if "Data" not in wb.sheetnames:
            raise ValueError(f"XLSX has no 'Data' sheet — found: {wb.sheetnames}")
        ws = wb["Data"]
        all_rows: list[tuple[Any, ...]] = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if len(all_rows) <= _ROW_DATA_START:
        raise ValueError(
            f"XLSX Data sheet has only {len(all_rows)} rows; "
            f"expected at least {_ROW_DATA_START + 1}"
        )

    series_id_row = all_rows[_ROW_SERIES_IDS]

    # Map column index → canonical metric name for mapped series only
    col_to_metric: dict[int, str] = {}
    for col_idx, cell in enumerate(series_id_row):
        if cell and cell in mappings:
            col_to_metric[col_idx] = mappings[cell]

    if not col_to_metric:
        logger.warning(
            "No series IDs in XLSX matched the mappings in metrics.yaml; output will be empty"
        )
        return []

    logger.info(
        "Found %d mapped series across %d data columns",
        len(col_to_metric),
        len(series_id_row),
    )

    output: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str]] = set()
    missing_count = 0
    duplicate_count = 0

    for row in all_rows[_ROW_DATA_START:]:
        date_cell = row[_COL_DATE] if len(row) > _COL_DATE else None
        institution_cell = row[_COL_INSTITUTION] if len(row) > _COL_INSTITUTION else None

        if not date_cell or not institution_cell:
            continue

        period = _to_quarter(date_cell)
        entity = str(institution_cell).strip()

        for col_idx, metric in col_to_metric.items():
            value = row[col_idx] if col_idx < len(row) else None

            if value is None:
                missing_count += 1
                logger.warning(
                    "Missing value: entity=%s metric=%r period=%s", entity, metric, period
                )

            dedup_key = (entity, metric, period)
            if dedup_key in seen:
                duplicate_count += 1
                logger.warning(
                    "Duplicate row skipped: entity=%s metric=%r period=%s",
                    entity,
                    metric,
                    period,
                )
                continue
            seen.add(dedup_key)

            output.append(
                {
                    "entity": entity,
                    "metric": metric,
                    "value": value,
                    "period": period,
                    "source": _SOURCE_ID,
                    "entity_type": "group" if entity in _GROUP_ENTITIES else "standalone",
                }
            )

    logger.info(
        "Parsed %d rows (%d missing values, %d duplicates skipped)",
        len(output),
        missing_count,
        duplicate_count,
    )
    return output


def _to_quarter(date_val: Any) -> str:
    """Convert a quarter-end date to a ``YYYY-QN`` string.

    Parameters
    ----------
    date_val:
        A ``datetime`` object or ISO-format string such as ``2024-09-30``.

    Returns
    -------
    str
        Quarter string, e.g. ``"2024-Q3"``.  Falls back to ``str(date_val)``
        if the value cannot be parsed.
    """
    if isinstance(date_val, datetime):
        q = (date_val.month - 1) // 3 + 1
        return f"{date_val.year}-Q{q}"
    try:
        dt = datetime.fromisoformat(str(date_val))
        q = (dt.month - 1) // 3 + 1
        return f"{dt.year}-Q{q}"
    except (ValueError, TypeError):
        logger.warning("Could not parse date value: %r", date_val)
        return str(date_val)


def write_csv(rows: list[dict[str, Any]], dest: Path) -> None:
    """Write canonical rows to a CSV file.

    Parameters
    ----------
    rows:
        Rows in canonical schema (entity, metric, value, period, source).
    dest:
        Destination file path.  Parent directories are created if absent.
    """
    dest.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["entity", "entity_type", "metric", "value", "period", "source"]
    with dest.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    logger.info("Wrote %d rows to %s", len(rows), dest)


def write_json(rows: list[dict[str, Any]], dest: Path) -> None:
    """Write canonical rows to a JSON file.

    Parameters
    ----------
    rows:
        Rows in canonical schema (entity, metric, value, period, source).
    dest:
        Destination file path.  Parent directories are created if absent.
    """
    dest.parent.mkdir(parents=True, exist_ok=True)
    with dest.open("w", encoding="utf-8") as fh:
        json.dump(rows, fh)
    logger.info("Wrote %d rows to %s", len(rows), dest)
